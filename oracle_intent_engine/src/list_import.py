"""
list_import.py
==============
List Import — CSV/Excel ingestion with field mapping and saved templates.

Workflow:
  1. User uploads CSV/Excel file via POST /api/import/upload
  2. Server parses headers and returns them alongside HubSpot field suggestions
  3. User maps columns → HubSpot fields (optionally saves template)
  4. POST /api/import/process runs ETL: parse → validate → DQE → stage
  5. Staged records flow into companies / company_contacts tables with status='staged'

POST-IMPORT COMPLETION (automatic):
  Company imports — every imported company is queued for the contact-enrichment
    pipeline (same Oracle/IT/finance title criteria as the lead enrichment
    workflow). unified_app launches the enrichment subprocess with the
    imported company ids right after the upload completes.

  Contact imports — each row is completed before saving:
    • email + LinkedIn  → ZeroBounce validates the email; saved with its status
                          (hard failures: invalid/spamtrap/abuse are NOT saved)
    • email only        → Apollo person-match fills the missing LinkedIn URL,
                          then ZeroBounce validates the email
    • name only         → Apollo person-match (name + company) finds the email
                          AND LinkedIn, then ZeroBounce validates
"""

import csv
import io
import json
import re
from typing import Optional

import openpyxl

import oracle_intent_engine.src.database as db
from oracle_intent_engine.src.data_quality import run_dqe_on_company, run_dqe_on_contact

# ── HubSpot field definitions (single source of truth) ───────────────────────

HS_COMPANY_FIELDS = [
    {"key": "name",                       "label": "Company Name",               "required": True},
    {"key": "domain",                     "label": "Website / Domain",           "required": False},
    {"key": "phone",                      "label": "Phone",                      "required": False},
    {"key": "industry",                   "label": "Industry",                   "required": False},
    {"key": "number_of_employees",        "label": "Number of Employees",        "required": False},
    {"key": "about_us",                   "label": "About Us",                   "required": False},
    {"key": "billing_street",             "label": "Billing Street",             "required": False},
    {"key": "billing_city",               "label": "Billing City",               "required": False},
    {"key": "billing_state",              "label": "Billing State",              "required": False},
    {"key": "billing_postal_code",        "label": "Billing Postal Code",        "required": False},
    {"key": "billing_country",            "label": "Billing Country",            "required": False},
    {"key": "duns_number",                "label": "DUNS Number",                "required": False},
    {"key": "holding_type",               "label": "Holding Type",               "required": False},
    {"key": "number_of_locations",        "label": "Number of Locations",        "required": False},
    {"key": "oracle_cloud_solutions",     "label": "Oracle Cloud Solutions",     "required": False},
    {"key": "oracle_on_premise_solutions","label": "Oracle On-Premise Solutions","required": False},
    {"key": "oracle_relationship_type",   "label": "Oracle Relationship Type",   "required": False},
    {"key": "oracle_support_end_date",    "label": "Oracle Support End Date",    "required": False},
    {"key": "oracle_version",             "label": "Oracle Version",             "required": False},
    {"key": "number_of_oracle_users",     "label": "Number of Oracle Users",     "required": False},
    {"key": "inoapps_account_manager",    "label": "Inoapps Account Manager",    "required": False},
    {"key": "inoapps_account_tier",       "label": "Inoapps Account Tier",       "required": False},
    {"key": "inoapps_relationship_type",  "label": "Inoapps Relationship Type",  "required": False},
    {"key": "inoapps_services_summary",   "label": "Inoapps Services Summary",   "required": False},
    {"key": "location",                   "label": "Location",                   "required": False},
    {"key": "size",                       "label": "Size / Revenue Band",        "required": False},
    {"key": "target_product",             "label": "Target Oracle Product",      "required": False},
]

HS_CONTACT_FIELDS = [
    {"key": "first_name",      "label": "First Name",       "required": True},
    {"key": "last_name",       "label": "Last Name",        "required": True},
    {"key": "email",           "label": "Email",            "required": False},
    {"key": "phone",           "label": "Phone",            "required": False},
    {"key": "mobile_phone",    "label": "Mobile Phone",     "required": False},
    {"key": "title",           "label": "Job Title",        "required": False},
    {"key": "job_function",    "label": "Job Function",     "required": False},
    {"key": "level",           "label": "Level / Seniority","required": False},
    {"key": "linkedin_url",    "label": "LinkedIn URL",     "required": False},
    {"key": "city",            "label": "City",             "required": False},
    {"key": "state",           "label": "State / Region",   "required": False},
    {"key": "country",         "label": "Country",          "required": False},
    {"key": "salutation",      "label": "Salutation",       "required": False},
    {"key": "suffix",          "label": "Suffix",           "required": False},
    {"key": "do_not_call",     "label": "Do Not Call",      "required": False},
    {"key": "do_not_email",    "label": "Do Not Email",     "required": False},
    {"key": "oracle_alignment", "label": "Oracle Alignment","required": False},
    {"key": "oracle_department","label": "Oracle Department","required": False},
    {"key": "oracle_team",     "label": "Oracle Team",      "required": False},
    {"key": "company_name",    "label": "Company Name",     "required": False},
]

_FIELD_LISTS = {
    "company":      HS_COMPANY_FIELDS,
    "contact":      HS_CONTACT_FIELDS,
}


# Explicit aliases: cleaned header string → field key.
# Covers shorthand / unconventional column names that the fuzzy scorer misses.
_ALIASES: dict[str, str] = {
    # Company fields
    "company":        "name",
    "companyname":    "name",
    "organisation":   "name",
    "organization":   "name",
    "org":            "name",
    "coname":         "name",
    "biz":            "name",
    "bizname":        "name",
    "accountname":    "name",
    "website":        "domain",
    "url":            "domain",
    "web":            "domain",
    "site":           "domain",
    "employees":      "number_of_employees",
    "headcount":      "number_of_employees",
    "empcount":       "number_of_employees",
    "numemployees":   "number_of_employees",
    "street":         "billing_street",
    "address":        "billing_street",
    "addr":           "billing_street",
    "city":           "billing_city",
    "state":          "billing_state",
    "zip":            "billing_postal_code",
    "zipcode":        "billing_postal_code",
    "postalcode":     "billing_postal_code",
    "postcode":       "billing_postal_code",
    "country":        "billing_country",
    "nation":         "billing_country",
    # Contact fields
    "fname":          "first_name",
    "firstname":      "first_name",
    "givenname":      "first_name",
    "forename":       "first_name",
    "lname":          "last_name",
    "lastname":       "last_name",
    "surname":        "last_name",
    "familyname":     "last_name",
    "jobtitle":       "title",
    "role":           "title",
    "position":       "title",
    "designation":    "title",
    "linkedin":       "linkedin_url",
    "linkedinprofile":"linkedin_url",
    "linkedinlink":   "linkedin_url",
    "mobile":         "mobile_phone",
    "cell":           "mobile_phone",
    "cellphone":      "mobile_phone",
    "mobilephone":    "mobile_phone",
    "emailaddress":   "email",
    "mail":           "email",
}


def _fuzzy_suggest(header: str, entity_type: str) -> Optional[str]:
    """Auto-suggest a HubSpot field key from a CSV column header.

    Resolution order:
      1. Exact alias match (covers shorthands like fname, org, zip)
      2. Exact cleaned string match against field key or label (score 100)
      3. Substring match against key or label (score 80 / 60)
    Only returns a suggestion when the winning score >= 60.
    """
    fields = _FIELD_LISTS.get(entity_type.lower(), [])
    if not fields:
        return None

    h = re.sub(r"[^a-z0-9]", "", header.lower())

    # 1. Alias lookup — always wins if present
    alias_key = _ALIASES.get(h)
    if alias_key and any(f["key"] == alias_key for f in fields):
        return alias_key

    # 2 & 3. Fuzzy scoring
    best, best_score = None, 0
    for f in fields:
        label_clean = re.sub(r"[^a-z0-9]", "", f["label"].lower())
        key_clean   = re.sub(r"[^a-z0-9]", "", f["key"].lower())
        score = 0
        if h == key_clean or h == label_clean:
            score = 100
        elif h in key_clean or key_clean in h:
            score = 80
        elif h in label_clean or label_clean in h:
            score = 60
        if score > best_score:
            best, best_score = f["key"], score
    return best if best_score >= 60 else None


# ── Mapping templates ─────────────────────────────────────────────────────────

def list_templates(entity_type: str = "") -> list:
    with db.db_cursor(commit=False) as cur:
        if entity_type:
            cur.execute(
                "SELECT * FROM import_mapping_templates WHERE entity_type=%s ORDER BY name",
                (entity_type,),
            )
        else:
            cur.execute("SELECT * FROM import_mapping_templates ORDER BY entity_type, name")
        return [dict(r) for r in cur.fetchall()]


def save_template(name: str, entity_type: str, mappings: dict, user_id: int = None) -> dict:
    with db.db_cursor() as cur:
        cur.execute(
            """INSERT INTO import_mapping_templates (name, entity_type, mappings, created_by)
               VALUES (%s,%s,%s,%s)
               ON CONFLICT (name, entity_type) DO UPDATE SET
                   mappings = EXCLUDED.mappings, updated_at = NOW()
               RETURNING *""",
            (name, entity_type, json.dumps(mappings), user_id),
        )
        return dict(cur.fetchone())


def delete_template(template_id: int) -> bool:
    with db.db_cursor() as cur:
        cur.execute(
            "DELETE FROM import_mapping_templates WHERE id=%s RETURNING id", (template_id,)
        )
        return cur.fetchone() is not None


# ── CSV parsing ───────────────────────────────────────────────────────────────

def _is_xlsx(content: bytes) -> bool:
    """Detect Excel .xlsx by its PK zip magic bytes."""
    return content[:4] == b"PK\x03\x04"


def _parse_xlsx_headers(content: bytes) -> list:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(c).strip() if c is not None else "" for c in row]
        break
    wb.close()
    return [h for h in headers if h]


def _parse_xlsx_rows(content: bytes) -> list:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return []
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    result = []
    for row in rows_iter:
        record = {}
        for col, val in zip(headers, row):
            if col:
                record[col] = str(val).strip() if val is not None else ""
        result.append(record)
    wb.close()
    return result


def parse_csv_headers(content: bytes, entity_type: str) -> dict:
    """
    Parse a CSV or Excel (.xlsx) file and return:
      headers      — list of {csv_header, suggested_field} for the frontend mapping UI
      record_count — number of data rows (excluding header)
      fields       — available HubSpot field definitions
    """
    et = entity_type.lower()
    if _is_xlsx(content):
        raw_headers = _parse_xlsx_headers(content)
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        record_count = max(0, ws.max_row - 1) if ws.max_row else 0
        wb.close()
    else:
        text = content.decode("utf-8-sig", errors="replace")
        lines = [l for l in io.StringIO(text)]
        record_count = max(0, len(lines) - 1)
        reader = csv.reader(io.StringIO(text))
        try:
            raw_headers = next(reader)
        except StopIteration:
            raw_headers = []

    if not raw_headers:
        return {"headers": [], "record_count": 0, "fields": _FIELD_LISTS.get(et, [])}

    headers = []
    for h in raw_headers:
        h = h.strip()
        if not h:
            continue
        suggested = _fuzzy_suggest(h, et)
        headers.append({"csv_header": h, "suggested_field": suggested or ""})

    return {
        "headers":      headers,
        "record_count": record_count,
        "fields":       _FIELD_LISTS.get(et, []),
    }


def process_import(
    content: bytes,
    entity_type: str,
    mappings: dict,
    batch_id: int,
    user_id: int = None,
    default_product: str = "",
    apollo_key: str = "",
    zerobounce_key: str = "",
) -> dict:
    """
    Run the ETL pipeline for an uploaded CSV.
    mappings = { "CSV Column Header": "hs_field_key", ... }

    apollo_key / zerobounce_key enable the contact completion flow
    (LinkedIn/email discovery + email validation) during contact imports.

    Returns { success_count, error_count, errors[], company_names[] }
    """
    if _is_xlsx(content):
        rows = _parse_xlsx_rows(content)
    else:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)

    success, errors = 0, []
    company_names: list = []   # companies imported this batch → auto-enrichment

    for idx, row in enumerate(rows, start=2):
        # Map CSV columns → HubSpot fields
        record: dict = {}
        for csv_col, hs_key in mappings.items():
            val = row.get(csv_col, "").strip()
            if val:
                record[hs_key] = val

        try:
            if entity_type == "company":
                if default_product and not record.get("target_product"):
                    record["target_product"] = default_product
                _import_company(record, user_id)
                if record.get("name"):
                    company_names.append(record["name"].strip())
            elif entity_type == "contact":
                _import_contact(record, user_id,
                                apollo_key=apollo_key, zerobounce_key=zerobounce_key)
            success += 1
        except Exception as e:
            errors.append({"row": idx, "error": str(e), "data": record})

    # Update batch stats
    with db.db_cursor() as cur:
        cur.execute(
            """UPDATE import_batches
               SET status='completed', success_count=%s, error_count=%s,
                   error_log=%s, completed_at=NOW()
               WHERE id=%s""",
            (success, len(errors), json.dumps(errors), batch_id),
        )

    return {"success_count": success, "error_count": len(errors),
            "errors": errors[:50], "company_names": company_names}


def _import_company(record: dict, user_id: int = None) -> None:
    name = record.get("name", "").strip()
    if not name:
        raise ValueError("Company name is required")

    # DQE check before insert
    issues = run_dqe_on_company(record)
    critical = [i for i in issues if i["severity"] == "critical"]
    if critical:
        raise ValueError(f"DQE: {critical[0]['message']}")

    with db.db_cursor() as cur:
        cur.execute(
            """INSERT INTO companies
                   (name, domain, industry, size, location, phone,
                    number_of_employees, billing_city, billing_country,
                    oracle_relationship_type, oracle_version,
                    inoapps_account_manager, inoapps_account_tier,
                    target_product, source, status)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'apollo','staged')
               ON CONFLICT (name) DO UPDATE SET
                   domain = COALESCE(EXCLUDED.domain, companies.domain),
                   industry = COALESCE(EXCLUDED.industry, companies.industry),
                   target_product = CASE WHEN EXCLUDED.target_product <> '' THEN EXCLUDED.target_product ELSE companies.target_product END,
                   last_updated = NOW()""",
            (
                name,
                record.get("domain") or None,   # nullable
                record.get("industry") or None,  # nullable
                record.get("size") or None,      # nullable
                record.get("location") or None,  # nullable
                record.get("phone") or "",
                record.get("number_of_employees") or None,  # nullable integer
                record.get("billing_city") or "",
                record.get("billing_country") or "",
                record.get("oracle_relationship_type") or "",
                record.get("oracle_version") or "",
                record.get("inoapps_account_manager") or "",
                record.get("inoapps_account_tier") or "",
                record.get("target_product") or "",
            ),
        )


# ZeroBounce statuses that mean the email must NOT be saved or contacted
_ZB_HARD_FAIL = {"invalid", "spamtrap", "abuse", "do_not_mail"}


def _complete_contact(record: dict, apollo_key: str, zerobounce_key: str) -> dict:
    """
    Fill missing email / LinkedIn via Apollo and validate the email via
    ZeroBounce, per the import completion rules (see module docstring).

    Returns {email, linkedin_url, title, email_source, validation_status,
             ready_for_outreach}.  Raises ValueError when the contact must
    not be saved (hard-failed email, or nothing found to contact them by).
    """
    from src.apollo_enrichment import apollo_person_match, _zb_validate_batch

    email        = str(record.get("email") or "").strip().lower()
    linkedin     = str(record.get("linkedin_url") or "").strip()
    title        = str(record.get("title") or "").strip()
    company_name = str(record.get("company_name") or "").strip()
    email_source = "import" if email else ""

    if email and not linkedin and apollo_key:
        # email only → Apollo fills the missing LinkedIn URL
        match = apollo_person_match(apollo_key, email=email)
        linkedin = match.get("linkedin_url") or linkedin
        title    = title or match.get("title", "")
    elif not email and apollo_key:
        # name only → Apollo finds email AND LinkedIn from name + company
        match = apollo_person_match(
            apollo_key,
            first_name=record.get("first_name", ""),
            last_name=record.get("last_name", ""),
            company_name=company_name or None,
        )
        if match.get("email"):
            email        = match["email"]
            email_source = "apollo"
        linkedin = linkedin or match.get("linkedin_url", "")
        title    = title or match.get("title", "")

    # ZeroBounce validation — only saved when the email is usable
    validation_status = ""
    if email:
        if zerobounce_key:
            zb = _zb_validate_batch([email], zerobounce_key)
            validation_status = zb.get(email, {}).get("status", "unknown")
            if validation_status in _ZB_HARD_FAIL:
                raise ValueError(
                    f"Email {email} failed ZeroBounce ({validation_status}) — contact not saved"
                )
        else:
            validation_status = "not_validated"
    elif not linkedin:
        raise ValueError("No email or LinkedIn found (Apollo had no match) — contact not saved")

    ready = bool(email and (
        validation_status == "valid"
        or (validation_status in ("catch-all", "catchall") and linkedin)
    ))
    return {
        "email":              email or None,
        "linkedin_url":       linkedin or None,
        "title":              title,
        "email_source":       email_source,
        "validation_status":  validation_status or None,
        "ready_for_outreach": ready,
    }


def _import_contact(record: dict, user_id: int = None,
                    apollo_key: str = "", zerobounce_key: str = "") -> None:
    first = record.get("first_name", "").strip()
    last  = record.get("last_name", "").strip()
    if not first or not last:
        raise ValueError("First name and last name are required")

    issues = run_dqe_on_contact(record)
    critical = [i for i in issues if i["severity"] == "critical"]
    if critical:
        raise ValueError(f"DQE: {critical[0]['message']}")

    # Complete the contact (Apollo fill-in + ZeroBounce validation) BEFORE saving
    completed = _complete_contact(record, apollo_key, zerobounce_key)

    # Find or create company; inherit its target_product for the contact
    company_name   = record.get("company_name", "").strip()
    company_id     = None
    target_product = ""
    if company_name:
        with db.db_cursor() as cur:
            cur.execute(
                "INSERT INTO companies (name, source, status) VALUES (%s,'apollo','staged') "
                "ON CONFLICT (name) DO UPDATE SET last_updated=NOW() "
                "RETURNING id, target_product",
                (company_name,),
            )
            row = cur.fetchone()
            company_id     = row["id"]
            target_product = row.get("target_product") or ""

    full_name = f"{first} {last}".strip()
    with db.db_cursor() as cur:
        cur.execute(
            """INSERT INTO company_contacts
                   (company_id, first_name, last_name, full_name, title, email, phone,
                    mobile_phone, linkedin_url, city, state, country,
                    job_function, level, oracle_alignment, oracle_department,
                    oracle_team, source, email_validation_status, email_source,
                    target_product, ready_for_outreach, status)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                       'import',%s,%s,%s,%s,'staged')
               ON CONFLICT DO NOTHING""",
            (
                company_id, first, last, full_name,
                completed["title"] or record.get("title"),
                completed["email"],
                record.get("phone"), record.get("mobile_phone"),
                completed["linkedin_url"], record.get("city"),
                record.get("state"), record.get("country"),
                record.get("job_function"), record.get("level"),
                record.get("oracle_alignment"), record.get("oracle_department"),
                record.get("oracle_team"),
                completed["validation_status"],
                completed["email_source"],
                target_product,
                completed["ready_for_outreach"],
            ),
        )


def create_batch(file_name: str, entity_type: str, record_count: int,
                 template_id: int = None, user_id: int = None) -> dict:
    with db.db_cursor() as cur:
        cur.execute(
            """INSERT INTO import_batches
                   (file_name, entity_type, mapping_template_id, record_count, created_by)
               VALUES (%s,%s,%s,%s,%s) RETURNING *""",
            (file_name, entity_type, template_id, record_count, user_id),
        )
        return dict(cur.fetchone())


def list_batches(limit: int = 50) -> list:
    with db.db_cursor(commit=False) as cur:
        cur.execute(
            """SELECT ib.*, imt.name AS template_name, u.email AS created_by_email
               FROM import_batches ib
               LEFT JOIN import_mapping_templates imt ON imt.id = ib.mapping_template_id
               LEFT JOIN users u ON u.id = ib.created_by
               ORDER BY ib.created_at DESC
               LIMIT %s""",
            (min(limit, 200),),
        )
        return [dict(r) for r in cur.fetchall()]
