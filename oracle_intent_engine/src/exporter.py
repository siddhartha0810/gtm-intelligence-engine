"""
Exports aggregated company data to CSV and Excel.
Output columns: company_name, oracle_product, phase, source, source_url,
                company_info, signal_count, confidence, evidence, detected_at
"""

import os
import csv
import pandas as pd
from datetime import datetime
from src.utils import get_logger
from src.phase_classifier import PHASE_LABELS
from src import lead_scorer
from src import database as db

logger = get_logger(__name__)

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output")


def _build_rows(companies: list[dict]) -> list[dict]:
    rows = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    for c in companies:
        company_info = ", ".join(filter(None, [
            c.get("industry", ""),
            c.get("size", ""),
            c.get("location", ""),
            c.get("website", ""),
        ]))

        # Prefer pre-fetched URL from DB query; fall back to scanning signals list
        source_url = c.get("source_url", "") or ""
        if not source_url:
            for sig in c.get("signals", []):
                url = sig.get("url", "")
                if url and url.startswith("http"):
                    source_url = url
                    break

        score = c.get("priority_score") or lead_scorer.calculate_priority_score(c)
        label = c.get("priority_label") or lead_scorer.get_priority_label(score)

        rows.append({
            "Priority Score":  score,
            "Priority":        label,
            "Company Name":    c.get("company_name", ""),
            "Oracle Product":  c.get("oracle_product", ""),
            "Phase":           PHASE_LABELS.get(c.get("phase", ""), c.get("phase", "")),
            "All Phases":      " / ".join(c.get("all_phases", [])),
            "Source(s)":       ", ".join(c.get("sources", [])),
            "Source URL":      source_url,
            "Company Info":    company_info or "N/A",
            "Location":        c.get("location", ""),
            "Signal Count":    c.get("signal_count", 0),
            "Evidence":        c.get("evidence", ""),
            "Detected At":     now,
        })
    return rows


def _build_contact_rows(companies: list[dict]) -> list[dict]:
    """
    Fetch matched contacts for the given companies only.
    All contacts remain stored in the DB; the Excel sheet shows only those
    belonging to companies included in this export.
    """
    rows = []
    try:
        names = [c.get("company_name", "") for c in companies if c.get("company_name")]
        contacts = db.get_contacts_for_company_names(names)
        for c in contacts:
            rows.append({
                "Company Name":   c.get("company_name", ""),
                "Company Domain": c.get("company_domain", ""),
                "First Name":     c.get("first_name", ""),
                "Last Name":      c.get("last_name", ""),
                "Email":          c.get("email", ""),
                "LinkedIn URL":   c.get("linkedin_url", ""),
            })
    except Exception as e:
        logger.error(f"Failed to build contact rows: {e}")
    return rows


def export_csv(companies: list[dict], filename: str = None) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if not filename:
        filename = f"oracle_intent_{ts}.csv"
    path = os.path.join(OUTPUT_DIR, filename)

    rows = _build_rows(companies)
    if not rows:
        logger.warning("No data to export.")
        return path

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    # Also export contacts as a separate CSV alongside
    contact_rows = _build_contact_rows(companies)
    if contact_rows:
        contacts_path = os.path.join(OUTPUT_DIR, f"oracle_contacts_{ts}.csv")
        with open(contacts_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=contact_rows[0].keys())
            writer.writeheader()
            writer.writerows(contact_rows)
        logger.info(f"Contacts CSV exported → {contacts_path} ({len(contact_rows)} rows)")

    logger.info(f"CSV exported → {path} ({len(rows)} rows)")
    return path


def export_excel(companies: list[dict], filename: str = None) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if not filename:
        filename = f"oracle_intent_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)

    rows = _build_rows(companies)
    contact_rows = _build_contact_rows(companies)

    if not rows and not contact_rows:
        logger.warning("No data to export.")
        return path

    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["Company Name"])

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Oracle Intent Signals", index=False)

        ws = writer.sheets["Oracle Intent Signals"]

        # Column widths
        # A=Priority Score, B=Priority, C=Company Name, D=Oracle Product,
        # E=Phase, F=All Phases, G=Source(s), H=Source URL,
        # I=Company Info, J=Location, K=Signal Count, L=Evidence, M=Detected At
        col_widths = {
            "A": 14, "B": 10, "C": 35, "D": 25,
            "E": 28, "F": 30, "G": 22, "H": 55,
            "I": 35, "J": 25, "K": 14, "L": 70, "M": 18,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Header styling
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Priority colour coding — column B (HOT/WARM/COLD)
        PRIORITY_HEX = {"HOT": "FFD7D7", "WARM": "FFE8C0", "COLD": "E8E8E8"}
        PRIORITY_TEXT = {"HOT": "CC0000", "WARM": "B35900", "COLD": "555555"}
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                val = str(cell.value or "")
                bg = PRIORITY_HEX.get(val, "FFFFFF")
                fg = PRIORITY_TEXT.get(val, "000000")
                cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                cell.font = Font(bold=True, color=fg)
                cell.alignment = Alignment(horizontal="center")

        # Phase colour coding — column E
        PHASE_HEX = {
            "Researching":            "D9E1F2",
            "Evaluating":             "BDD7EE",
            "Budgeting / Approving":  "FFE699",
            "Hiring for Oracle":      "C6EFCE",
            "Implementing":           "FFC7CE",
            "Post Go-Live / Support": "DDEBF7",
        }
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                phase_val = str(cell.value or "")
                hex_col = PHASE_HEX.get(phase_val, "FFFFFF")
                cell.fill = PatternFill(start_color=hex_col, end_color=hex_col, fill_type="solid")

        # Hyperlink styling for Source URL column (H)
        hyperlink_font = Font(color="0563C1", underline="single")
        for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
            for cell in row:
                url = str(cell.value or "")
                if url.startswith("http"):
                    cell.hyperlink = url
                    cell.font = hyperlink_font
                    cell.value = url

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # ── Matched Contacts sheet ──────────────────────────────────────────
        if contact_rows:
            df_contacts = pd.DataFrame(contact_rows)
            df_contacts.to_excel(writer, sheet_name="Matched Contacts", index=False)
            wc = writer.sheets["Matched Contacts"]

            contact_col_widths = {"A": 35, "B": 28, "C": 18, "D": 18, "E": 36, "F": 55}
            for col, width in contact_col_widths.items():
                wc.column_dimensions[col].width = width

            for cell in wc[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Hyperlink LinkedIn URLs
            li_col = 6  # column F
            for row in wc.iter_rows(min_row=2, min_col=li_col, max_col=li_col):
                for cell in row:
                    url = str(cell.value or "")
                    if url.startswith("http"):
                        cell.hyperlink = url
                        cell.font = Font(color="0563C1", underline="single")

            wc.freeze_panes = "A2"
            wc.auto_filter.ref = wc.dimensions

    logger.info(f"Excel exported → {path} ({len(rows)} rows, {len(contact_rows)} contacts)")
    return path
