"""
pipeline.py
===========
MAIN ENTRY POINT for the Lead Enrichment Engine.
Orchestrates all 7 pipeline stages end-to-end for a batch of leads.

PURPOSE:
  Takes a raw list of leads (names, company, title — no email) and produces
  a fully-enriched, ZeroBounce-validated list with emails and LinkedIn URLs
  that the sales team can import directly into their outreach tool.

HOW IT FITS IN THE SYSTEM:
  Invoked by enrichment_worker.py as:
    python -m src.pipeline <path_to_xlsx> --restart|--resume

  Also callable directly from the CLI for local testing.

  Data flow:
    Input Excel/CSV
      → Stage 1: clean + deduplicate
      → Stage 2: resolve company domains (8 parallel workers)
      → Stage 3: Apollo bulk_match + Apify fallback (10 parallel workers)
      → Stage 4: ZeroBounce validate vendor emails
      → Stage 5: email pattern engine predicts missing emails
      → Stage 6: ZeroBounce validate predicted emails
      → Stage 7: scoring — set ready_for_outreach = True/False
      → Output CSV + Excel write-back

KEY CLASSES/FUNCTIONS:
  run_pipeline()            — main orchestrator: calls stages in order
  _checkpoint_path()        — returns the .pkl checkpoint file path
  _save_checkpoint()        — saves DataFrame + stage name to disk
  _load_checkpoint()        — loads a checkpoint if --resume was passed
  _print_stage_summary()    — prints lead counts + credit usage after each stage
  _invalidate_suspect_domains() — evicts stale cached domains where all
                                  ZeroBounce-validated predictions failed
  _propagate_email_domains()    — fills missing domains for same-company leads
                                  once any lead in the group has a confirmed domain

CHECKPOINT / RESUME:
  Each stage saves a .pkl checkpoint after completing.
  --resume skips stages that already completed.
  --restart wipes the checkpoint and starts from Stage 1.
  This prevents losing hours of work if a run is interrupted.

EXCEL WRITE-BACK:
  The original .xlsx file is updated in-place using openpyxl.
  Columns added: email, linkedin_url, email_source, email_validation_status,
  ready_for_outreach (green = True, red = False), confidence_score.
  Header row: dark blue background, white bold text.

DEPENDENCIES:
  - orchestrator.py    : Stage 3 (vendor enrichment)
  - domain_resolver.py : Stage 2
  - zerobounce_client.py : Stages 4 + 6
  - email_pattern_engine.py : Stage 5
  - database.py        : caches domain lookups + Apollo results
  - pg_master.py       : contacts_master (read-only Salesforce CRM export)

Usage (from inside the lead_enrichment_engine folder):
  python -m src.pipeline "path/to/your/file.xlsx" --restart
  python -m src.pipeline "path/to/your/file.xlsx" --resume
  python -m src.pipeline                            (uses input/leads.csv)

Flags:
  --restart   discard any saved checkpoint and start fresh (use for new files)
  --resume    skip the resume prompt and continue from checkpoint automatically

Pipeline stages (in order):
  1. cleaned           — normalize columns, deduplicate rows
  2. domain_resolved   — find company website domains
  3. enriched          — call Apollo/Apify to get emails + LinkedIn URLs
  4. vendor_validated  — ZeroBounce validates Apollo/Apify emails
  5. email_predicted   — guess emails for remaining leads using name patterns
  6. pred_validated    — ZeroBounce validates predicted emails
  7. scored            — mark each lead yes/no for outreach

Output:
  output/final_outreach_ready.csv  — enriched leads
  output/audit_log.csv             — row counts per stage
  output/vendor_performance.csv    — Apollo hit/miss stats
  [original .xlsx file]            — results written back with new columns
"""

import math
import os
import sys
import time

import pandas as pd

from .audit import AuditLog
from .checkpoint import clear, load, save, should_skip
from .cleaner import clean_leads
from .config import (
    DB_PATH, DOMAIN_LOOKUP, INPUT_LEADS, OUTPUT_AUDIT, OUTPUT_FINAL, SUPPRESSION_LIST,
    PG_CONNECTION_STRING, PG_INPUT_TABLE, PG_OUTPUT_TABLE,
    PG_MASTER_CONNECTION_STRING,
)
from .database import get_db, init_db
from .domain_resolver import MAX_WORKERS as DOMAIN_WORKERS
from .domain_resolver import resolve_domains
from .email_pattern_engine import predict_missing_emails, collapse_prediction_candidates
from .orchestrator import MAX_WORKERS as ENRICH_WORKERS
from .orchestrator import enrich, get_apollo_credits
from .scoring import mark_outreach_ready
from .zerobounce_client import (
    BATCH_THRESHOLD, POLL_INTERVAL,
    _fmt_credits, get_credits, validate_emails,
)

# ── Constants ──────────────────────────────────────────────────────────────

# Email sources that are validated in Stage 4 (vendor emails)
VENDOR_SOURCES = ["input", "apollo", "apify", "zoominfo"]

# Columns written to the final output CSV — in display order
FINAL_COLUMNS = [
    "lead_id", "first_name", "last_name", "company", "domain",
    "email", "email_source", "email_validation_status", "email_validation_sub_status",
    "email_prediction_pattern", "email_prediction_confidence",
    "linkedin_url", "linkedin_source", "job_title",
    "ready_for_outreach", "failure_reason",
]

# Columns this pipeline adds to the Excel file — stripped before re-reading
# on subsequent runs so the pipeline always starts clean
_ADDED_COLS = {
    "Domain", "Email", "Email Source", "Email Status",
    "LinkedIn URL", "Ready for Outreach", "Failure Reason",
}


# ── Time Formatting Helpers ────────────────────────────────────────────────

def _fmt_dur(seconds: float) -> str:
    """Convert seconds to a human-readable duration string."""
    if seconds < 60:
        return f"{seconds:.0f}s"
    m, s = divmod(int(seconds), 60)
    if m < 60:
        return f"{m}m {s:02d}s"
    h, m = divmod(m, 60)
    return f"{h}h {m:02d}m"


def _estimate_runtime(df: pd.DataFrame, last_stage) -> dict:
    """
    Estimate how long each remaining stage will take.

    Based on typical API latencies:
      Domain resolution : 8 workers, ~4s per batch of 8 companies
      Vendor enrichment : 10 workers, ~4s per batch of 10 leads
      ZeroBounce batch  : ~70s per 200 emails (small batch)
      ZeroBounce bulk   : ~300s + 6s per 1,000 emails (large batch)
    """
    n            = len(df)
    n_no_domain  = int((df["domain"].astype(str).str.strip() == "").sum())
    n_needs      = int(
        (df["email"].astype(str).str.strip().eq("")
         | df.get("linkedin_url", pd.Series("", index=df.index)).astype(str).str.strip().eq("")).sum()
    )
    n_has_email  = int((df["email"].astype(str).str.strip() != "").sum())
    n_pred_est   = max(int(n_needs * 0.25), 0)   # rough guess: ~25% of gaps filled by prediction

    def _zb_secs(count):
        if count == 0: return 0
        if count <= BATCH_THRESHOLD:
            return max(math.ceil(count / 200) * 70, 5)
        return 300 + math.ceil(count / 1000) * 6

    fast = max(math.ceil(n / 10_000), 1)
    estimates = {
        "cleaned":          fast,
        "domain_resolved":  max(math.ceil(n_no_domain / DOMAIN_WORKERS) * 4, 2) if n_no_domain else 2,
        "enriched":         max(math.ceil(n_needs / ENRICH_WORKERS) * 4, 2)     if n_needs   else 2,
        "vendor_validated": _zb_secs(n_has_email),
        "email_predicted":  fast,
        "pred_validated":   _zb_secs(n_pred_est),
        "scored":           fast,
    }

    # Zero out stages already completed (resumed run)
    if last_stage:
        for stage in list(estimates):
            if should_skip(stage, last_stage):
                estimates[stage] = 0

    return estimates


def _print_eta(estimates: dict, last_stage) -> None:
    """Print the estimated runtime table to the terminal."""
    total = sum(estimates.values())
    print(f"  Estimated runtime")
    print(f"  {'-'*52}")
    labels = {
        "cleaned":          "Clean & deduplicate",
        "domain_resolved":  "Resolve domains",
        "enriched":         "Vendor enrichment",
        "vendor_validated": "Validate vendor emails (ZeroBounce)",
        "email_predicted":  "Predict missing emails",
        "pred_validated":   "Validate predicted emails (ZeroBounce)",
        "scored":           "Score & suppress",
    }
    for stage, secs in estimates.items():
        label = labels.get(stage, stage)
        tag   = "(skipped)" if secs == 0 else f"~{_fmt_dur(secs)}"
        print(f"    {label:<40} {tag}")
    print(f"  {'-'*52}")
    print(f"  Total estimated time          ~{_fmt_dur(total)}")
    print()


# ── Stage Logging ──────────────────────────────────────────────────────────

def _log(audit: AuditLog, df: pd.DataFrame, stage: str, elapsed: float) -> None:
    """Log a stage completion to the audit log and print a status line."""
    audit.log_stage_summary(stage, df)
    emails   = int((df["email"].astype(str).str.strip() != "").sum())
    linkedin = int((df["linkedin_url"].astype(str).str.strip() != "").sum()) if "linkedin_url" in df.columns else 0
    print(f"  OK {stage:<38} emails={emails}  linkedin={linkedin}  [{_fmt_dur(elapsed)}]")


def _load_csv(path: str) -> pd.DataFrame:
    """Load the input CSV file with helpful error messages if it's missing or empty."""
    try:
        df = pd.read_csv(path)
    except FileNotFoundError:
        print(f"\n  ERROR: Input file not found: {path}")
        print(f"  Create the file or pass a different path:")
        print(f"    python -m src.pipeline path/to/your/file.csv\n")
        sys.exit(1)
    except Exception as exc:
        print(f"\n  ERROR reading {path}: {exc}\n")
        sys.exit(1)
    if df.empty:
        print(f"\n  ERROR: {path} is empty.\n")
        sys.exit(1)
    return df


# ── Excel Write-Back ───────────────────────────────────────────────────────

def _write_back_xlsx(enriched_df: pd.DataFrame, source_path: str, dest_path: str) -> None:
    """
    Merge enriched columns back into the original Excel file and save.

    source_path: the original xlsx to read original columns from
    dest_path:   where to save the merged result (same as source_path normally)

    The function:
      1. Reads the original xlsx (stripping any previously-added enriched columns)
      2. Matches leads by (first_name, last_name, company) — case-insensitive
      3. Merges enriched columns alongside original columns
      4. Applies formatting: dark blue headers, green/red ready_for_outreach cells
      5. Saves back to dest_path
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Read original file, stripping any enriched columns from previous runs
    orig = pd.read_excel(source_path)
    orig = orig[[c for c in orig.columns if c not in _ADDED_COLS]]
    orig_col_count = len(orig.columns)

    # Build slim enriched DataFrame with display-friendly column names
    enrich_cols = ["first_name", "last_name", "company", "domain", "email",
                   "email_source", "email_validation_status", "linkedin_url",
                   "ready_for_outreach", "failure_reason"]
    slim = enriched_df[enrich_cols].copy()
    slim.columns = ["First Name", "Last Name", "_co_norm", "Domain", "Email",
                    "Email Source", "Email Status", "LinkedIn URL",
                    "Ready for Outreach", "Failure Reason"]

    # Create lowercase match keys for joining
    for frame in (orig, slim):
        frame["_fn"] = frame["First Name"].astype(str).str.strip().str.lower()
        frame["_ln"] = frame["Last Name"].astype(str).str.strip().str.lower()

    orig["_co"] = orig.iloc[:, 3].astype(str).str.strip().str.lower()   # 4th col = company
    slim["_co"] = slim["_co_norm"].astype(str).str.strip().str.lower()

    # Left join: keep all original rows, attach enriched columns where matched
    merged = orig.merge(
        slim.drop(columns=["First Name", "Last Name", "_co_norm"]),
        on=["_fn", "_ln", "_co"], how="left"
    ).drop(columns=["_fn", "_ln", "_co"])

    # Write merged data to Excel
    merged.to_excel(dest_path, index=False, sheet_name="Sheet1")

    # ── Apply formatting using openpyxl ────────────────────────────────────
    wb = load_workbook(dest_path)
    ws = wb.active

    total_cols = ws.max_column
    total_rows = ws.max_row

    thin   = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    data_font = Font(name="Arial", size=10)

    # Header row — dark blue for original columns, lighter blue for new columns
    for col in range(1, total_cols + 1):
        cell      = ws.cell(row=1, column=col)
        is_new    = col > orig_col_count
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", start_color="2E75B6" if is_new else "1F4E79")
        cell.alignment = center
        cell.border    = border

    # Find the "Ready for Outreach" column index for color coding
    ready_col = next(
        (c for c in range(1, total_cols + 1)
         if ws.cell(row=1, column=c).value == "Ready for Outreach"),
        None,
    )

    # Data rows — apply font, border, and alignment
    for row in range(2, total_rows + 1):
        for col in range(1, total_cols + 1):
            cell           = ws.cell(row=row, column=col)
            cell.font      = data_font
            cell.border    = border
            cell.alignment = left

        # Color-code the Ready for Outreach cell: green=yes, red=no
        if ready_col:
            cell = ws.cell(row=row, column=ready_col)
            val  = str(cell.value or "").lower()
            cell.alignment = center
            if val == "yes":
                cell.fill = PatternFill("solid", start_color="C6EFCE")
                cell.font = Font(name="Arial", size=10, bold=True, color="276221")
            elif val == "no":
                cell.fill = PatternFill("solid", start_color="FFC7CE")
                cell.font = Font(name="Arial", size=10, bold=True, color="9C0006")

    # Set column widths for readability
    col_widths = {
        "Domain": 24, "Email": 34, "Email Source": 14,
        "Email Status": 14, "LinkedIn URL": 48,
        "Ready for Outreach": 18, "Failure Reason": 35,
    }
    for col in range(1, total_cols + 1):
        header = ws.cell(row=1, column=col).value
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(str(header), 20)

    # Freeze the top row so headers stay visible when scrolling
    ws.freeze_panes = "A2"

    wb.save(dest_path)
    print(f"  Excel updated        : {dest_path}")


# ── Main Pipeline ──────────────────────────────────────────────────────────

def _invalidate_suspect_domains(df: pd.DataFrame, lookup_path: str) -> None:
    """
    Self-healing cache: after Stage 6, find companies where EVERY predicted
    email came back invalid and the domain was auto-resolved (not manually set).

    Those domains are removed from domain_lookup.csv so the next run re-resolves
    them from APIs with MX validation — instead of re-using a wrong cached value.

    High-confidence (manual) entries are never touched.
    """
    from pathlib import Path
    from .utils import normalize_company

    path = Path(lookup_path)
    if not path.exists():
        return

    predicted = df[df["email_source"].astype(str) == "predicted"]
    if predicted.empty:
        return

    # Companies where ALL prediction attempts returned invalid status
    suspect_norms: set = set()
    for norm, grp in predicted.groupby("company_normalized"):
        statuses = grp["email_validation_status"].astype(str).str.strip()
        if statuses.isin(["invalid", ""]).all():
            suspect_norms.add(norm)

    if not suspect_norms:
        return

    lookup = pd.read_csv(path)
    lookup["_norm"] = lookup["company"].apply(normalize_company)

    # Only evict auto-resolved entries — never remove manual high-confidence rows
    evict_mask = (
        lookup["_norm"].isin(suspect_norms) &
        lookup["confidence"].astype(str).str.lower().ne("high")
    )

    if not evict_mask.any():
        return

    evicted = lookup.loc[evict_mask, ["company", "domain"]].to_dict("records")
    lookup   = lookup[~evict_mask].drop(columns=["_norm"])
    lookup.to_csv(path, index=False)

    print(f"\n    domain cache: evicted {len(evicted)} suspect domain(s) — will re-resolve next run:")
    for row in evicted:
        print(f"      {row['company']:35s} was [{row['domain']}]")


def _propagate_email_domains(df: pd.DataFrame) -> pd.DataFrame:
    """
    After vendor validation (Stage 4), use confirmed email domains from leads
    that already have a ZeroBounce-valid email to fix the domain for other
    leads at the same company that still need an email.

    Example: Apollo found adam@mhs-pa.org (valid) for Milton Hershey School,
    but the domain resolver set domain=mhskids.org for Aliena Anderson (same
    company). This corrects Aliena's domain to mhs-pa.org so the prediction
    engine generates aanderson@mhs-pa.org instead of aanderson@mhskids.org.
    """
    df = df.copy()

    has_valid = (
        df["email"].astype(str).str.strip().ne("") &
        df["email_validation_status"].astype(str).eq("valid")
    )
    if not has_valid.any():
        return df

    # Extract email domain from every lead with a confirmed-valid email
    confirmed = df.loc[has_valid, ["company_normalized", "email"]].copy()
    confirmed["_edomain"] = (
        confirmed["email"].astype(str).str.split("@").str[-1].str.lower().str.strip()
    )

    # Pick the most common email domain per company
    company_email_domain: dict = {}
    for norm, grp in confirmed.groupby("company_normalized"):
        top = grp["_edomain"].value_counts()
        if not top.empty:
            company_email_domain[norm] = top.index[0]

    if not company_email_domain:
        return df

    # Apply to leads still missing an email where a better domain is known
    needs_email = df["email"].astype(str).str.strip().eq("")
    changes = 0
    for idx, row in df.loc[needs_email].iterrows():
        norm    = str(row.get("company_normalized", "")).strip()
        edomain = company_email_domain.get(norm)
        if not edomain:
            continue
        current = str(row.get("domain", "")).strip().lower()
        if edomain != current:
            df.at[idx, "domain"] = edomain
            changes += 1
            if changes <= 8:
                company = str(row.get("company", "")).strip()
                print(f"    domain fix [{company}]: {current or '(none)'} -> {edomain}")

    if changes:
        print(f"    email-domain propagation: {changes} lead(s) corrected")

    return df


def _validate_config() -> bool:
    """
    Pre-flight check: warn about missing or incomplete API keys before any stage runs.
    Returns True if at least one enrichment vendor is configured, False otherwise.
    Prints warnings for anything that will silently degrade results.
    """
    from .config import (
        APOLLO_API_KEY, ZEROBOUNCE_API_KEY,
        APIFY_TOKEN, APIFY_LINKEDIN_ACTOR_ID, APIFY_EMAIL_ACTOR_ID,
    )

    warnings: list = []
    infos:    list = []
    has_apollo = bool(APOLLO_API_KEY)
    has_apify  = bool(APIFY_TOKEN and APIFY_LINKEDIN_ACTOR_ID and APIFY_EMAIL_ACTOR_ID)

    # Apollo
    if not APOLLO_API_KEY:
        warnings.append("APOLLO_API_KEY not set — Apollo enrichment disabled")
    else:
        infos.append("Apollo        : configured")

    # Apify
    if not APIFY_TOKEN:
        infos.append("Apify         : not configured (no APIFY_TOKEN)")
    elif not APIFY_LINKEDIN_ACTOR_ID or not APIFY_EMAIL_ACTOR_ID:
        warnings.append(
            "APIFY_TOKEN is set but actor IDs are missing — Apify disabled "
            "(need both APIFY_LINKEDIN_ACTOR_ID and APIFY_EMAIL_ACTOR_ID)"
        )
    else:
        infos.append("Apify         : configured")

    # ZeroBounce
    if not ZEROBOUNCE_API_KEY:
        warnings.append(
            "ZEROBOUNCE_API_KEY not set — emails will be marked 'not_validated' "
            "and NO leads will be ready for outreach"
        )
    else:
        infos.append("ZeroBounce    : configured")

    # No enrichment vendor at all
    if not has_apollo and not has_apify:
        warnings.append(
            "No enrichment vendor configured — pipeline will find NO new emails. "
            "Set at least APOLLO_API_KEY in .env"
        )

    print(f"\n  API Configuration")
    print(f"  {'-'*52}")
    for info in infos:
        print(f"    {info}")
    if warnings:
        print()
        for w in warnings:
            print(f"    !! WARNING: {w}")

    return has_apollo or has_apify


def _show_credits() -> None:
    """Print current ZeroBounce and Apollo credit balances and exit."""
    from .orchestrator import get_apollo_credits

    zb  = get_credits()
    apo = get_apollo_credits()

    print(f"\n  Credit Balances")
    print(f"  {'='*40}")

    print(f"\n  ZeroBounce")
    print(f"  {'-'*40}")
    print(f"  Balance          : {_fmt_credits(zb)}")

    print(f"\n  Apollo")
    print(f"  {'-'*40}")
    src = apo.get("source", "")
    if src == "api":
        rem  = apo.get("credits_remaining")
        used = apo.get("credits_used")
        lim  = apo.get("credits_limit")
        if rem  is not None: print(f"  Credits remaining: {rem:,}")
        if used is not None: print(f"  Credits used     : {used:,}")
        if lim  is not None: print(f"  Credits limit    : {lim:,}")
    elif src == "master_key_required":
        print(f"  Balance          : unavailable (standard key — master API key required)")
    elif src == "no_key":
        print(f"  Balance          : not checked (no APOLLO_API_KEY in .env)")
    else:
        print(f"  Balance          : could not fetch")
    print()


def main() -> None:
    """
    Run the full lead enrichment pipeline.

    Reads input → runs all 7 stages → writes output CSV → writes back to xlsx.
    Supports checkpoint/resume so a crashed run can be continued.
    """
    os.makedirs("output", exist_ok=True)
    pipeline_start = time.time()
    audit   = AuditLog()

    # ── Database Init ──────────────────────────────────────────────────────
    db = init_db(DB_PATH)
    db.purge_expired()
    imported = db.import_csv(DOMAIN_LOOKUP)
    stats    = db.stats()

    # Initialize PostgreSQL master store (always Inoapps-Data-DB)
    pg_master = None
    try:
        from .pg_master import init_pg_master
        pg_master = init_pg_master()
        pg_ms = pg_master.master_stats()
        print(f"\n  Master store (PG) : {pg_ms['total']:,} total  |  {pg_ms['with_email']:,} with email  |  {pg_ms['valid']:,} validated  |  {pg_ms['ready']:,} ready")
    except Exception as exc:
        print(f"\n  WARNING: Could not connect to Inoapps-Data-DB master store: {exc}")
        pg_master = None

    print(f"\n  Database : {stats['domains']} domains  |  {stats['patterns']} patterns  |  {stats['cached_leads']} cached leads  |  {stats.get('master_leads', 0)} contacts (ZB-validated)", end="")
    if imported:
        print(f"  (+{imported} imported from CSV)", end="")
    print()
    args    = sys.argv[1:]
    restart = "--restart" in args
    resume  = "--resume"  in args

    if "--credits" in args:
        _show_credits()
        return

    input_file = next((a for a in args if not a.startswith("--")), INPUT_LEADS)

    # ── Auto-convert Excel input to CSV ────────────────────────────────────
    # If the user passes a .xlsx file, convert it to CSV for processing.
    # We strip any previously-added enriched columns so the pipeline always
    # starts from the original data, regardless of what's in the file.
    source_xlsx = None
    if input_file.lower().endswith((".xlsx", ".xlsm", ".xls")):
        source_xlsx = input_file
        csv_path    = INPUT_LEADS
        raw         = pd.read_excel(source_xlsx)
        raw         = raw[[c for c in raw.columns if c not in _ADDED_COLS]]
        raw.to_csv(csv_path, index=False)
        print(f"\n  Converted {os.path.basename(source_xlsx)} -> {csv_path}  ({len(raw)} rows, {len(raw.columns)} original columns)")
        input_file = csv_path

    # ── Checkpoint / Resume ────────────────────────────────────────────────
    # Load any saved checkpoint from a previous run
    df, last_stage = load()

    if df is not None and last_stage and not restart:
        if not resume:
            print(f"\n  Checkpoint found (last stage: '{last_stage}')")
            if input("  Resume? [y/n]: ").strip().lower() != "y":
                clear()
                df, last_stage = None, None
        else:
            print(f"\n  RESUMING from checkpoint (last completed: '{last_stage}')")
    elif restart:
        clear()
        df, last_stage = None, None

    if df is None:
        explicit_file = any(a for a in args if not a.startswith("--"))
        if PG_CONNECTION_STRING and not explicit_file:
            from .pg_connector import load_leads as _pg_load
            try:
                pg_df = _pg_load(PG_CONNECTION_STRING, PG_INPUT_TABLE)
            except RuntimeError as exc:
                print(f"\n  ERROR (Postgres input): {exc}")
                print(f"  Falling back to CSV: {input_file}\n")
                pg_df = None
            if pg_df is not None:
                df = pg_df
                print(f"\n  Input   : PostgreSQL  table='{PG_INPUT_TABLE}'  ({len(df)} rows)")
            else:
                print(f"\n  Postgres table '{PG_INPUT_TABLE}' is empty — falling back to CSV: {input_file}")
                df = _load_csv(input_file)
        else:
            df = _load_csv(input_file)
        last_stage = None

    # ── API config pre-flight ──────────────────────────────────────────────
    _validate_config()

    # ── Stage Runner ───────────────────────────────────────────────────────
    def ckpt(frame, stage_name):
        """Save a mid-run checkpoint (called every 50 leads during enrichment)."""
        save(frame, stage_name)

    stage_times: dict = {}

    def run(stage, fn, *args, **kwargs):
        """
        Run a single pipeline stage.

        If the stage was already completed (resume mode), skip it and show
        the current counts. Otherwise, run the function, time it, log it,
        and save a checkpoint.
        """
        nonlocal df

        # Skip if this stage was already done in a previous run
        if should_skip(stage, last_stage):
            emails   = int((df["email"].astype(str).str.strip() != "").sum())
            linkedin = int((df["linkedin_url"].astype(str).str.strip() != "").sum()) if "linkedin_url" in df.columns else 0
            print(f"  -- {stage:<38} emails={emails}  linkedin={linkedin}  [skipped]")
            stage_times[stage] = 0
            audit.log_stage_summary(stage, df)
            return df

        t0      = time.time()
        result  = fn(*args, **kwargs)
        elapsed = time.time() - t0
        stage_times[stage] = elapsed
        _log(audit, result, stage, elapsed)
        save(result, stage)
        return result

    def _persist_to_master(frame, label=""):
        """
        Write enriched results to the PostgreSQL master store.
        Called after every key stage so partial runs still accumulate data.
        """
        if not pg_master:
            print(f"  WARNING: pg_master not initialised — results not persisted to master")
            return
        records = frame.to_dict("records")
        tag = f" [{label}]" if label else ""
        try:
            pg_master.upsert_master_leads(records)
            ms = pg_master.master_stats()
            print(f"  Master (PG){tag}: {ms['total']:,} total  |  {ms['with_email']:,} with email  |  {ms['valid']:,} validated  |  {ms['ready']:,} ready")
        except Exception as exc:
            print(f"  WARNING: PG master upsert failed{tag}: {exc}")

    # ── Pipeline Header ────────────────────────────────────────────────────
    print(f"\n  Lead Enrichment Pipeline")
    print(f"  {'='*52}")
    print(f"  Input : {input_file}  ({len(df)} rows)")

    # ── ZeroBounce Opening Balance ─────────────────────────────────────────
    # Show credit balance upfront so the user knows if they have enough credits
    opening_balance = get_credits()
    existing_emails = int((df["email"].astype(str).str.strip() != "").sum()) if "email" in df.columns else 0
    leads_no_email  = len(df) - existing_emails

    print(f"\n  ZeroBounce Credits")
    print(f"  {'-'*52}")
    if opening_balance is not None:
        print(f"  Balance now          : {_fmt_credits(opening_balance)}")
        print(f"  Emails already in CSV: {existing_emails:,}  (will be validated)")
        print(f"  Leads missing email  : {leads_no_email:,}  (enrichment + prediction may add more)")
        if opening_balance < existing_emails:
            print(f"\n  !! WARNING: balance ({opening_balance:,}) < known emails to validate ({existing_emails:,}).")
            print(f"  !! Top up at zerobounce.net before continuing.")
    else:
        print(f"  Balance now          : not checked (no ZEROBOUNCE_API_KEY in .env)")
        print(f"  Emails will be marked 'not_validated' until a key is set.")
    print()

    # ── STAGE 1: Clean ─────────────────────────────────────────────────────
    # Run clean first (it's fast) so we have real data for the ETA estimate
    df = run("cleaned", clean_leads, df)

    # Show estimated runtime based on actual data shape
    estimates = _estimate_runtime(df, last_stage)
    _print_eta(estimates, last_stage)

    apollo_credits_before: dict = {}
    apollo_credits_after:  dict = {}

    # ── STAGE 2: Domain Resolution ─────────────────────────────────────────
    df = run("domain_resolved", resolve_domains, df, DOMAIN_LOOKUP)

    # ── STAGE 3: Vendor Enrichment (Apollo → Apify) ────────────────────────
    apollo_credits_before = get_apollo_credits()
    df = run("enriched", enrich, df, checkpoint_fn=ckpt)
    apollo_credits_after  = get_apollo_credits()

    # Log Apollo credit consumption for this run so the UI can show per-step burn
    try:
        import sys, os as _os
        _root = _os.path.dirname(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))
        if _root not in sys.path:
            sys.path.insert(0, _root)
        from oracle_intent_engine.src.database_sqlite import log_apollo_credits
        _run_id = str(run_id) if run_id else "standalone"
        _before = apollo_credits_before.get("consumed_credits") if isinstance(apollo_credits_before, dict) else None
        _after  = apollo_credits_after.get("consumed_credits")  if isinstance(apollo_credits_after,  dict) else None
        log_apollo_credits(_run_id, "stage_3_enrichment", _before, _after)
    except Exception as _e:
        pass  # credit logging is best-effort — never block the pipeline

    # ── STAGE 4: Validate Vendor Emails ────────────────────────────────────
    # Validate emails found by Apollo, Apify, or present in the original input
    df = run("vendor_validated", validate_emails, df, source_filter=VENDOR_SOURCES)

    # Persist Apollo emails + ZeroBounce validation to master immediately.
    # If the run crashes before Stage 7, this data is not lost.
    if not should_skip("vendor_validated", last_stage):
        _persist_to_master(df, "after stage 4")

    # ── Cache enriched results after validation ────────────────────────────
    # Persists email + ZeroBounce status to DB so re-runs skip Apollo entirely.
    if not should_skip("email_predicted", last_stage):
        has_email = df["email"].astype(str).str.strip().ne("")
        if has_email.any():
            db.cache_leads(df[has_email].to_dict("records"))

    # ── STAGE 3b: Re-enrich leads whose input email was invalid ──────────────
    # If a lead arrived with an email from the input CSV but ZeroBounce flagged
    # it as invalid, clear it and ask Apollo to find a replacement before the
    # prediction engine runs. This covers the edge case: input has email + a
    # known invalid address (e.g. old email, typo, dead mailbox).
    if not should_skip("email_predicted", last_stage):
        invalid_input_mask = (
            df["email_validation_status"].astype(str).str.strip().eq("invalid")
            & df["email_source"].astype(str).str.strip().eq("input")
        )
        if invalid_input_mask.any():
            n = int(invalid_input_mask.sum())
            print(f"\n  STAGE 3b: {n} lead(s) had invalid input email — sending to Apollo for replacement")
            # Clear invalid email so orchestrator includes them in the "needs" list
            for col in ["email", "email_source", "email_validation_status", "email_validation_sub_status"]:
                df.loc[invalid_input_mask, col] = ""
            # Evict from enrichment cache so the re-run doesn't restore the bad email
            bad_ids = [str(df.at[idx, "lead_id"]) for idx in df.index[invalid_input_mask]]
            db.evict_leads(bad_ids)
            df = run("re_enriched", enrich, df, checkpoint_fn=ckpt)
            df = run("re_validated", validate_emails, df, source_filter=["apollo", "apify"])

    # ── Domain propagation: use confirmed email domains to fix same-company leads ──
    # Runs between Stage 4 and Stage 5 so the prediction engine gets the right domain.
    # Only fires if Stage 4 is not being skipped (resume mode).
    if not should_skip("email_predicted", last_stage):
        df = _propagate_email_domains(df)
        save(df, "vendor_validated")   # update checkpoint with corrected domains

    # ── STAGE 5: Predict Missing Emails ────────────────────────────────────
    # For leads still without an email, try to guess it from naming patterns
    df = run("email_predicted", predict_missing_emails, df)

    # ── PRE-FLIGHT GATE: Test 25-row sample before full Stage 6 spend ─────────
    # Validates a small sample first to check authentication rate.
    # If the rate is below 70%, the domain patterns are unreliable and we skip
    # the full run — saving the bulk of ZeroBounce credits.
    _GATE_SAMPLE   = 25
    _GATE_MIN_RATE = 0.70
    _predicted_pool = df[
        (df["email_source"].astype(str) == "predicted") &
        (df["email"].astype(str).str.strip() != "")
    ]

    _run_stage6 = True
    if not should_skip("pred_validated", last_stage) and len(_predicted_pool) > _GATE_SAMPLE:
        print(f"\n  Pre-flight gate      : validating {_GATE_SAMPLE}-row sample before full Stage 6")
        _sample = _predicted_pool.sample(_GATE_SAMPLE, random_state=42)
        _sample_validated = validate_emails(_sample.copy(), source_filter=["predicted"])
        _valid_statuses = {"valid", "catch-all"}
        _valid_n = int(_sample_validated["email_validation_status"].isin(_valid_statuses).sum())
        _rate    = _valid_n / _GATE_SAMPLE
        print(f"  Pre-flight result    : {_valid_n}/{_GATE_SAMPLE} valid ({_rate:.0%}) — threshold {_GATE_MIN_RATE:.0%}")
        if _rate < _GATE_MIN_RATE:
            print(f"  !! Gate FAILED — skipping Stage 6 to save ZeroBounce credits.")
            print(f"  !! Review domain patterns or check email_prediction_pattern distribution.")
            _run_stage6 = False
        else:
            # Copy sample validation results back into main df so Stage 6 doesn't re-bill them
            df.update(_sample_validated[["email_validation_status", "email_validation_sub_status"]])
            print(f"  Pre-flight PASSED — proceeding with full Stage 6")

    # ── STAGE 6: Validate Predicted Emails ─────────────────────────────────
    # Run ZeroBounce again, but only on the predicted emails
    if _run_stage6:
        df = run("pred_validated", validate_emails, df, source_filter=["predicted"])
    else:
        save(df, "pred_validated")  # checkpoint so resume works

    # Collapse multi-candidate global-pattern rows back to 1 row per lead,
    # keeping the first ZeroBounce-valid prediction and dropping the rest.
    df = collapse_prediction_candidates(df)
    save(df, "pred_validated")

    # Persist predicted + validated emails to master immediately after Stage 6.
    if not should_skip("pred_validated", last_stage):
        _persist_to_master(df, "after stage 6")

    # Emit per-pattern validation breakdown so the UI Stage 05 card can show
    # which formats succeeded and how many leads each produced.
    _pred = df[df["email_source"].astype(str).str.strip() == "predicted"]
    if not _pred.empty:
        _valid     = _pred[_pred["email_validation_status"].astype(str).str.strip() == "valid"]
        _total     = len(_pred)
        _valid_n   = len(_valid)
        _v_parts   = [f"{pat}:{int(cnt)}" for pat, cnt in _valid["email_prediction_pattern"].value_counts().items()]
        _inv_parts = [f"{pat}:{int(cnt)}" for pat, cnt in
                      _pred[_pred["email_validation_status"].astype(str).str.strip() == "invalid"]
                      ["email_prediction_pattern"].value_counts().items()]
        print(f"    predict valid patterns  : {' | '.join(_v_parts) if _v_parts else 'none'}  ({_valid_n}/{_total} valid)")
        if _inv_parts:
            print(f"    predict invalid patterns: {' | '.join(_inv_parts)}")

    # Self-healing: evict domains from cache where every prediction was invalid.
    # Next run will re-resolve these companies from APIs with MX validation.
    _invalidate_suspect_domains(df, DOMAIN_LOOKUP)

    # Sync DB → CSV so domain_lookup.csv reflects the current DB state for human review.
    db.export_csv(DOMAIN_LOOKUP)

    # ── STAGE 7: Score & Suppress ──────────────────────────────────────────
    # Mark each lead yes/no based on email validation status and suppression list
    df = run("scored", mark_outreach_ready, df, SUPPRESSION_LIST)

    # ── Write Outputs ──────────────────────────────────────────────────────

    # Ensure all expected columns exist before writing
    for col in FINAL_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    # Write enriched CSV
    df[FINAL_COLUMNS].to_csv(OUTPUT_FINAL, index=False)
    audit.save(OUTPUT_AUDIT)

    # Final persist — writes scored ready_for_outreach status to both stores
    _persist_to_master(df)

    # Write enriched results to Postgres (if configured)
    if PG_CONNECTION_STRING:
        from .pg_connector import save_results as _pg_save
        try:
            rows_saved = _pg_save(df[FINAL_COLUMNS], PG_CONNECTION_STRING, PG_OUTPUT_TABLE)
            print(f"  Postgres output      : {rows_saved} rows → table '{PG_OUTPUT_TABLE}'")
        except RuntimeError as exc:
            print(f"\n  WARNING (Postgres output): {exc}")
            print(f"  Results are still available in {OUTPUT_FINAL}")

    # Write results back into the original Excel file (automatically)
    if source_xlsx:
        try:
            _write_back_xlsx(df, source_xlsx, source_xlsx)
        except PermissionError:
            # File is open in Excel — write to a timestamped sibling file instead
            ts  = time.strftime("%Y%m%d_%H%M%S")
            alt = source_xlsx.replace(".xlsx", f"_enriched_{ts}.xlsx")
            print(f"  !! {os.path.basename(source_xlsx)} is open in Excel — writing to:")
            print(f"     {alt}")
            _write_back_xlsx(df, source_xlsx, alt)

    # Delete checkpoint files — run is complete
    clear()

    total_elapsed = time.time() - pipeline_start
    ready         = int((df["ready_for_outreach"] == "yes").sum())
    closing_balance = get_credits()

    # ── Final Summary ──────────────────────────────────────────────────────
    print(f"\n  {'='*52}")
    print(f"  RESULTS")
    print(f"  {'='*52}")
    print(f"  Ready for outreach   : {ready:,} / {len(df):,} leads")
    print(f"  Output CSV           : {OUTPUT_FINAL}")
    if source_xlsx:
        print(f"  Output Excel         : {source_xlsx}")
    print(f"  Audit log            : {OUTPUT_AUDIT}")
    print(f"  Vendor performance   : output/vendor_performance.csv")
    closing_stats = db.stats()
    print(f"  DB (pipeline.db)     : {closing_stats['domains']} domains  |  {closing_stats['patterns']} patterns  |  {closing_stats['cached_leads']} cached leads")

    # Time breakdown per stage
    print(f"\n  Time Summary")
    print(f"  {'-'*52}")
    stage_labels = {
        "cleaned":          "Clean & deduplicate",
        "domain_resolved":  "Resolve domains",
        "enriched":         "Vendor enrichment",
        "vendor_validated": "Validate vendor emails",
        "email_predicted":  "Predict missing emails",
        "pred_validated":   "Validate predicted emails",
        "scored":           "Score & suppress",
    }
    for stage, actual in stage_times.items():
        est   = estimates.get(stage, 0)
        label = stage_labels.get(stage, stage)
        if actual == 0:
            print(f"    {label:<38} skipped")
        else:
            diff     = actual - est
            diff_str = f"  ({'+' if diff >= 0 else ''}{_fmt_dur(abs(diff))} {'over' if diff >= 0 else 'under'} est.)"
            print(f"    {label:<38} {_fmt_dur(actual):<8}{diff_str}")
    print(f"  {'-'*52}")
    print(f"  Total (wall clock)   : {_fmt_dur(total_elapsed)}")

    # ── Credit Summary ─────────────────────────────────────────────────────
    print(f"\n  Credit Summary")
    print(f"  {'='*52}")

    # ZeroBounce
    print(f"\n  ZeroBounce")
    print(f"  {'-'*52}")
    if opening_balance is not None and closing_balance is not None:
        used = opening_balance - closing_balance
        print(f"  Credits at start     : {opening_balance:,}")
        print(f"  Credits at end       : {closing_balance:,}")
        print(f"  Credits used         : {used:,}")
    elif closing_balance is not None:
        print(f"  Credits remaining    : {closing_balance:,}")
    else:
        print(f"  Credits remaining    : not available (no API key)")

    # Apollo
    from .orchestrator import _fmt_apollo_credits
    print(f"\n  Apollo (end of run)")
    print(f"  {'-'*52}")
    apo_src = apollo_credits_after.get("source", "")
    print(f"  Credits              : {_fmt_apollo_credits(apollo_credits_after)}")
    if apo_src == "api":
        rm = apollo_credits_after.get("rate_limit_minute")
        rh = apollo_credits_after.get("rate_limit_hour")
        if rm: print(f"  Rate limit / minute  : {rm:,}")
        if rh: print(f"  Rate limit / hour    : {rh:,}")
    print()


if __name__ == "__main__":
    main()
