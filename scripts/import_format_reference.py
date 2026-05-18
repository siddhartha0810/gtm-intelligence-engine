"""
import_format_reference.py
==========================
One-time (and idempotent) import of COMPANY_FORMAT_ANALYSIS.xlsx into the
email_patterns table in PostgreSQL.

The file contains up to 16 ranked email-format slots per company/domain,
each with a format code (e.g. "flast", "first.last") and a sample_count
derived from ~280k validated contacts.  This gives the prediction engine
high-confidence domain-specific patterns before any live enrichment runs.

Usage:
  python scripts/import_format_reference.py [path/to/COMPANY_FORMAT_ANALYSIS.xlsx]

If no path is given, the script looks for the file at:
  - C:/Users/<user>/Downloads/COMPANY_FORMAT_ANALYSIS.xlsx
  - ./COMPANY_FORMAT_ANALYSIS.xlsx

Environment variables (or .env in the project root):
  DB_HOST, DB_PORT, DB_NAME, DB_USER, DB_PASSWORD
  or ORACLE_PG_DSN=postgresql://user:pass@host:5432/dbname

The import is idempotent: repeated runs UPDATE the sample_count to the
higher of the existing vs imported value, so live data is never overwritten
with a smaller reference count.
"""

import os
import sys
import time
from pathlib import Path

# ── Try to load .env from project root ────────────────────────────────────────
_ROOT = Path(__file__).parent.parent
try:
    from dotenv import load_dotenv
    _env = _ROOT / "oracle_intent_engine" / ".env"
    if _env.exists():
        load_dotenv(_env)
        print(f"Loaded env from {_env}")
except ImportError:
    pass   # dotenv not installed — rely on real env vars

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    sys.exit("ERROR: psycopg2-binary is required. Install with: pip install psycopg2-binary")

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl is required. Install with: pip install openpyxl")


# ── Format code normalisation ─────────────────────────────────────────────────
# Maps codes from the Excel (Primary/Secondary Format Code column) to the
# internal pattern names used by the prediction engine.
# Codes that cannot be predicted deterministically are mapped to None (skipped).

FORMAT_CODE_MAP: dict[str, str | None] = {
    "flast":               "flast",       # j + smith  -> jsmith
    "first.last":          "first.last",  # john + smith -> john.smith
    "f.last":              "f.last",      # j.smith
    "first":               "first",       # john
    "firstlast":           "firstlast",   # johnsmith
    "last.first":          "last.first",  # smith.john
    "last.f":              "last.f",      # smith.j
    "lastf":               "lastf",       # smithj
    "first.l":             "first.l",     # john.s
    "firstl":              "firstl",      # johns  (first name + last initial)
    "last":                "last",        # smith  (last name only)
    # ── skip these — no deterministic formula ─────────────────────────────────
    "flast + numeric suffix":  None,
    "f...last / custom":       None,
    "first...last / custom":   None,
    "other / unmatched":       None,
}

# Format slots: each is 8 columns wide starting at col index 14 (0-based)
# Slot layout: [code, formula, generic_ex, domain_ex, desc, count, share_pct, samples]
_N_SLOTS     = 16
_SLOT_START  = 14   # index of "Primary Format Code"
_SLOT_WIDTH  = 8    # columns per slot

# Column indices within a slot
_IDX_CODE  = 0  # relative: format code  (e.g. "flast")
_IDX_COUNT = 5  # relative: sample count (integer, how many emails confirmed this format)

# Absolute column indices for domain (col 1) and company (col 0)
_COL_DOMAIN  = 1
_COL_COMPANY = 0


def _build_dsn() -> str:
    dsn = os.environ.get("ORACLE_PG_DSN", "").strip()
    if dsn:
        return dsn
    host     = os.environ.get("DB_HOST", "localhost")
    port     = os.environ.get("DB_PORT", "5432")
    dbname   = os.environ.get("DB_NAME", "oracle_intent")
    user     = os.environ.get("DB_USER", "postgres")
    password = os.environ.get("DB_PASSWORD", "")
    return f"host={host} port={port} dbname={dbname} user={user} password={password}"


def _find_file() -> Path:
    """Locate the Excel file: argv[1], Downloads, or cwd."""
    if len(sys.argv) > 1:
        p = Path(sys.argv[1])
        if not p.exists():
            sys.exit(f"ERROR: File not found: {p}")
        return p

    candidates = [
        Path.home() / "Downloads" / "COMPANY_FORMAT_ANALYSIS.xlsx",
        Path.home() / "OneDrive" / "Downloads" / "COMPANY_FORMAT_ANALYSIS.xlsx",
        Path.cwd() / "COMPANY_FORMAT_ANALYSIS.xlsx",
    ]
    for c in candidates:
        if c.exists():
            return c
    sys.exit(
        "ERROR: Could not find COMPANY_FORMAT_ANALYSIS.xlsx.\n"
        "Pass the path as the first argument:\n"
        "  python scripts/import_format_reference.py path/to/file.xlsx"
    )


def _ensure_table(cur):
    """Create email_patterns table if it doesn't exist (mirrors database.py DDL)."""
    cur.execute("""
        CREATE TABLE IF NOT EXISTS email_patterns (
            domain        TEXT NOT NULL,
            pattern       TEXT NOT NULL,
            sample_count  INTEGER NOT NULL DEFAULT 1,
            last_seen_at  TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            PRIMARY KEY (domain, pattern)
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_ep_domain ON email_patterns(domain)")


def _parse_rows(ws) -> list[tuple[str, str, int]]:
    """
    Read the sheet and yield (domain, pattern, sample_count) tuples.
    Skips rows with no domain, no usable format, or count=0.
    """
    tuples: list[tuple[str, str, int]] = []
    seen: set[tuple[str, str]] = set()   # dedup (domain, pattern) within the file

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        domain = str(row[_COL_DOMAIN] or "").strip().lower()
        if not domain or domain == "none":
            continue

        for slot in range(_N_SLOTS):
            code_col  = _SLOT_START + (slot * _SLOT_WIDTH) + _IDX_CODE
            count_col = _SLOT_START + (slot * _SLOT_WIDTH) + _IDX_COUNT

            if code_col >= len(row):
                break

            raw_code = str(row[code_col] or "").strip().lower() if row[code_col] else ""
            if not raw_code:
                break   # no more slots for this row

            pattern = FORMAT_CODE_MAP.get(raw_code)
            if not pattern:
                continue   # unmappable code — skip

            raw_count = row[count_col] if count_col < len(row) else None
            try:
                count = int(raw_count) if raw_count is not None else 1
            except (ValueError, TypeError):
                count = 1
            count = max(count, 1)   # treat 0 as 1 (domain still uses this pattern)

            key = (domain, pattern)
            if key in seen:
                continue
            seen.add(key)
            tuples.append((domain, pattern, count))

    return tuples


def run():
    xlsx_path = _find_file()
    print(f"File: {xlsx_path}")
    print("Opening workbook (read-only)…")
    t0 = time.time()

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    sheet_name = "DETAILED COMPANY FORMAT GUIDE"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    print(f"Sheet '{sheet_name}': {ws.max_row:,} rows × {ws.max_column} cols")

    print("Parsing format data…")
    rows = _parse_rows(ws)
    print(f"  -> {len(rows):,} (domain, pattern, count) records extracted  "
          f"({time.time()-t0:.1f}s)")

    if not rows:
        print("Nothing to import.")
        return

    dsn = _build_dsn()
    print(f"\nConnecting to PostgreSQL…")
    conn = psycopg2.connect(dsn)
    conn.autocommit = False
    cur = conn.cursor()

    _ensure_table(cur)
    conn.commit()

    print(f"Upserting {len(rows):,} rows into email_patterns…")
    t1 = time.time()

    # Only 3 columns — last_seen_at uses the column DEFAULT (NOW())
    # ON CONFLICT updates last_seen_at explicitly so repeated imports refresh it
    upsert_sql = """
        INSERT INTO email_patterns (domain, pattern, sample_count)
        VALUES %s
        ON CONFLICT (domain, pattern) DO UPDATE SET
            sample_count = GREATEST(email_patterns.sample_count, EXCLUDED.sample_count),
            last_seen_at = NOW()
    """

    # Batch in chunks of 2000 to avoid huge single transactions
    CHUNK = 2_000
    total_upserted = 0
    for i in range(0, len(rows), CHUNK):
        batch = rows[i:i + CHUNK]
        vals  = [(d, p, c) for d, p, c in batch]
        psycopg2.extras.execute_values(cur, upsert_sql, vals, page_size=500)
        conn.commit()
        total_upserted += len(batch)
        pct = total_upserted / len(rows) * 100
        print(f"  {total_upserted:>7,} / {len(rows):,}  ({pct:.0f}%)  "
              f"{time.time()-t1:.1f}s", end="\r", flush=True)

    print()

    # Stats
    cur.execute("SELECT COUNT(*) FROM email_patterns")
    total_in_db = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT domain) FROM email_patterns")
    domains_in_db = cur.fetchone()[0]
    cur.execute("SELECT pattern, COUNT(*) AS n FROM email_patterns GROUP BY pattern ORDER BY n DESC")
    pattern_dist = cur.fetchall()

    conn.close()

    elapsed = time.time() - t0
    print(f"\n{'='*55}")
    print(f"  Import complete in {elapsed:.1f}s")
    print(f"  Rows upserted : {total_upserted:,}")
    print(f"  Total in DB   : {total_in_db:,} rows across {domains_in_db:,} domains")
    print(f"\n  Pattern distribution:")
    for pat, n in pattern_dist:
        print(f"    {pat:<20} {n:>8,} domains")
    print(f"{'='*55}")
    print("\nDone. The prediction engine will now use this reference data.")


if __name__ == "__main__":
    run()
