#!/usr/bin/env python3
"""
import_contacts.py
==================
Load the VALIDATED slice of the Salesforce contact export into the DB, grouped
by company, with each contact's email + LinkedIn.

What this build does differently:
  * Keeps only ZeroBounce-valid contacts (ready_for_outreach) — the ~222K we
    actually outreach to, not the full 276K.
  * Normalizes company identity: curated proper names + industry + HQ for the
    top accounts (company_reference.py), algorithmic title-case for the rest.
    Several domains collapse into one parent (jpmorgan.com/chase.com →
    JPMorgan Chase), so the company count reflects real companies.
  * Derives each company's location: curated HQ where known, else the majority
    mailing location of its contacts.

Idempotent: clears only prior source='contacts_master' rows before loading.
Backend-agnostic (SQLite + Postgres).

Usage:
    python import_contacts.py [/path/to/ALL_CONTACTS_CONSOLIDATED.xlsx] [--limit=N]
"""
from __future__ import annotations

import re
import sys
import uuid
from collections import Counter, defaultdict
from pathlib import Path

BASE = Path(__file__).parent
ORACLE = BASE / "intent_engine"
if str(ORACLE) not in sys.path:
    sys.path.insert(0, str(ORACLE))

DEFAULT_PATH = "/Users/sid/Desktop/ALL_CONTACTS_CONSOLIDATED.xlsx"
SOURCE_TAG = "contacts_master"
BATCH = 5000

_SUFFIX = {"inc", "llc", "ltd", "corp", "co", "company", "group", "holdings",
           "plc", "sa", "ag", "gmbh", "the", "and"}


def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _truthy(v) -> bool:
    return _s(v).lower() in ("1", "yes", "true")


def normalize_name(raw: str) -> str:
    """Best-effort cleanup for long-tail company names: title-case each token,
    preserve short acronyms. (Mangled single-tokens like 'Wellsfargo' are fixed
    by the curated reference; this handles the ordinary cases.)"""
    if not raw:
        return ""
    words = [w for w in re.split(r"[\s_\-]+", raw.strip()) if w]
    out = []
    for w in words:
        if w.isupper() and len(w) <= 4:
            out.append(w)                      # keep acronyms (IBM, AT&T-ish)
        else:
            out.append(w[:1].upper() + w[1:].lower())
    return " ".join(out).strip()


def _fmt_loc(city: str, state: str, country: str) -> str:
    return ", ".join(p for p in (city, state, country) if p)


def main() -> None:
    path = DEFAULT_PATH
    limit = None
    for a in sys.argv[1:]:
        if a.startswith("--limit="):
            limit = int(a.split("=", 1)[1])
        elif not a.startswith("--"):
            path = a

    import openpyxl
    from src import database as db, company_reference as ref

    print(f"Source: {path}")
    db.init_db()

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    idx = {h: i for i, h in enumerate(header)}

    def g(row, col):
        i = idx.get(col)
        return _s(row[i]) if i is not None and i < len(row) else ""

    print("Streaming (valid contacts only) + normalizing…")
    # company_key → aggregate; company_key is the canonical proper name
    companies: dict[str, dict] = {}
    contacts: list[dict] = []
    seen: set[tuple] = set()
    skipped_invalid = skipped_no_co = dupes = 0

    for n, row in enumerate(it):
        if limit and n >= limit:
            break
        if not _truthy(g(row, "ZB_Valid_Email")):
            skipped_invalid += 1
            continue
        email = (g(row, "Validated_Email") or g(row, "Email")).lower()
        if not email:
            skipped_invalid += 1
            continue
        domain = g(row, "Domain").lower()
        raw_company = g(row, "Existing_Company") or g(row, "New_Company")

        r = ref.lookup(domain)
        if r:
            name, industry, hq_city, hq_state, hq_country = r
        else:
            name = normalize_name(raw_company)
            industry, hq_city, hq_state, hq_country = "", "", "", ""
        if not name:
            skipped_no_co += 1
            continue

        key = (name, email)
        if key in seen:
            dupes += 1
            continue
        seen.add(key)

        c_city, c_state, c_country = g(row, "MailingCity"), g(row, "MailingState"), g(row, "MailingCountry")

        agg = companies.get(name)
        if agg is None:
            agg = {"domain": Counter(), "industry": industry,
                   "hq": _fmt_loc(hq_city, hq_state, hq_country),
                   "loc": Counter()}
            companies[name] = agg
        if domain:
            agg["domain"][domain] += 1
        if not agg["hq"]:  # non-curated: learn location from contacts
            cl = _fmt_loc(c_city, c_state, c_country)
            if cl:
                agg["loc"][cl] += 1

        contacts.append({
            "company": name,
            "first_name": g(row, "FirstName"),
            "last_name": g(row, "LastName"),
            "title": g(row, "Title"),
            "email": email,
            "linkedin": g(row, "LinkedIn_URL_Enriched") or g(row, "LinkedIn_URL__c"),
            "seniority": g(row, "Management_Level__c"),
            "job_function": g(row, "Job_Function__c"),
            "domain": domain,
            "phone": g(row, "Phone"),
            "mobile": g(row, "MobilePhone"),
            "city": c_city, "state": c_state, "country": c_country,
            "dnc": _truthy(g(row, "DoNotCall")),
            "dne": _truthy(g(row, "HasOptedOutOfEmail")) or _truthy(g(row, "Overall_Opt_Out__c")),
            "moved": _truthy(g(row, "ZI_Person_has_Moved__c")),
        })
        if (n + 1) % 50000 == 0:
            print(f"  …{n + 1:,} rows scanned")

    print(f"Kept {len(contacts):,} valid contacts across {len(companies):,} companies "
          f"(skipped {skipped_invalid:,} invalid, {skipped_no_co:,} no-company, {dupes:,} dupes)")

    # Resolve each company's domain + location
    for name, agg in companies.items():
        agg["domain_str"] = agg["domain"].most_common(1)[0][0] if agg["domain"] else ""
        agg["location"] = agg["hq"] or (agg["loc"].most_common(1)[0][0] if agg["loc"] else "")

    # ── Upsert companies (name, domain, industry, location) ───────────────────
    print("Upserting companies…")
    with db.db_cursor() as cur:
        cur.executemany(
            "INSERT INTO companies (name, domain, industry, location) VALUES (%s, %s, %s, %s) "
            "ON CONFLICT (name) DO UPDATE SET "
            "  domain   = COALESCE(NULLIF(EXCLUDED.domain,''),   companies.domain), "
            "  industry = COALESCE(NULLIF(EXCLUDED.industry,''), companies.industry), "
            "  location = COALESCE(NULLIF(EXCLUDED.location,''), companies.location)",
            [(name, a["domain_str"], a["industry"], a["location"]) for name, a in companies.items()],
        )
    with db.db_cursor(commit=False) as cur:
        cur.execute("SELECT id, name FROM companies")
        name_to_id = {r["name"]: r["id"] for r in cur.fetchall()}

    # ── Clear prior corpus rows, batch-insert contacts ────────────────────────
    print(f"Clearing prior source='{SOURCE_TAG}' contacts…")
    with db.db_cursor() as cur:
        cur.execute("DELETE FROM company_contacts WHERE source = %s", (SOURCE_TAG,))

    print("Batch-inserting contacts…")
    cols = ("company_id, first_name, last_name, full_name, title, email, linkedin_url, "
            "seniority, source, email_validation_status, ready_for_outreach, unique_key, "
            "domain, phone, mobile_phone, city, state, country, do_not_call, do_not_email, "
            "person_has_moved, job_function, level")
    # ON CONFLICT DO NOTHING: consolidating aliased domains into one company can
    # surface the same person (same email/LinkedIn) twice — keep the first.
    sql = f"INSERT INTO company_contacts ({cols}) VALUES ({', '.join(['%s'] * 23)}) ON CONFLICT DO NOTHING"

    inserted = 0
    batch: list[tuple] = []
    for c in contacts:
        cid = name_to_id.get(c["company"])
        if not cid:
            continue
        full = f"{c['first_name']} {c['last_name']}".strip()
        batch.append((
            cid, c["first_name"], c["last_name"], full, c["title"], c["email"], c["linkedin"],
            c["seniority"], SOURCE_TAG, "valid", True, uuid.uuid4().hex,
            c["domain"], c["phone"], c["mobile"], c["city"], c["state"], c["country"],
            bool(c["dnc"]), bool(c["dne"]), bool(c["moved"]),
            c["job_function"], c["seniority"],
        ))
        if len(batch) >= BATCH:
            with db.db_cursor() as cur:
                cur.executemany(sql, batch)
            inserted += len(batch); batch = []
            if inserted % 40000 == 0:
                print(f"  …{inserted:,} inserted")
    if batch:
        with db.db_cursor() as cur:
            cur.executemany(sql, batch)
        inserted += len(batch)

    # ── Refresh counts, prune companies left with zero contacts ───────────────
    print("Updating counts + pruning empty companies…")
    with db.db_cursor() as cur:
        cur.execute("""
            UPDATE companies SET contact_count = (
                SELECT COUNT(*) FROM company_contacts
                WHERE company_contacts.company_id = companies.id AND email != '')
        """)
    with db.db_cursor() as cur:
        cur.execute("DELETE FROM companies WHERE contact_count = 0 AND "
                    "id NOT IN (SELECT DISTINCT company_id FROM intent_signals WHERE company_id IS NOT NULL)")

    with db.db_cursor(commit=False) as cur:
        cur.execute("SELECT COUNT(*) AS n FROM company_contacts WHERE source = %s", (SOURCE_TAG,))
        actual = cur.fetchone()["n"]
    print(f"\nDONE — {actual:,} valid contacts stored ({inserted - actual:,} duplicate rows skipped).")
    _sample(db)


def _sample(db) -> None:
    with db.db_cursor(commit=False) as cur:
        cur.execute("SELECT COUNT(*) n FROM companies")
        print(f"Companies now: {cur.fetchone()['n']:,}")
        cur.execute("""
            SELECT c.name, c.industry, c.location, COUNT(cc.id) AS n
            FROM companies c JOIN company_contacts cc ON cc.company_id = c.id
            WHERE cc.source = %s
            GROUP BY c.id ORDER BY n DESC LIMIT 10
        """, (SOURCE_TAG,))
        print("\nTop companies (normalized name · industry · location · contacts):")
        for r in cur.fetchall():
            print(f"   {r['n']:5,}  {(r['name'] or '')[:26]:26} | {(r['industry'] or '—')[:22]:22} | {(r['location'] or '—')[:26]:26}")


if __name__ == "__main__":
    main()
