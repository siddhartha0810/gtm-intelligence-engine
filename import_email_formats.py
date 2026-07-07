#!/usr/bin/env python3
"""
import_email_formats.py
=======================
Load the validated per-domain email formats into the prediction engine.

Source: COMPANY-EMAIL-FORMAT-REFERENCE-GUIDE.xlsx (sheet "DETAILED COMPANY
FORMAT GUIDE") — one row per domain×format, derived from the 280K validated
corpus, with the format code and how many validated emails matched it.

Writes to two tables in one pass:

  email_patterns (domain, pattern, sample_count) — the lean operational table
  email_pattern_engine reads via load_domain_patterns() to build a predicted
  email. Only format codes with a buildable formula (see PATTERNS in
  email_pattern_engine.py) are loaded here — "Other / unmatched" and custom
  multi-dot codes have no template to build from and are skipped.

  company_email_formats (domain, format_rank, ...) — the FULL guide, every
  row, including the unbuildable codes. This is the read model for the
  Prediction Engine UI: every company's format(s) should be visible and
  searchable even when the engine can't auto-generate an address from it —
  "no template, but here's what we know" is still useful to a rep.

Usage:
    python import_email_formats.py [/path/to/GUIDE.xlsx]
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

BASE = Path(__file__).parent
ORACLE = BASE / "oracle_intent_engine"
if str(ORACLE) not in sys.path:
    sys.path.insert(0, str(ORACLE))

DEFAULT_PATH = "/Users/sid/Downloads/COMPANY-EMAIL-FORMAT-REFERENCE-GUIDE.xlsx"
SHEET = "DETAILED COMPANY FORMAT GUIDE"
BATCH = 5000

# Codes email_pattern_engine.PATTERNS can build. Guide codes outside this set
# (e.g. "Other / unmatched", "f...last / custom") are unusable for generation —
# skipped in email_patterns, but still kept (flagged) in company_email_formats.
_ENGINE_CODES = {
    "first.last", "firstlast", "flast", "first_last", "f.last", "first.l",
    "last.first", "first", "lastf", "last.f", "firstl", "last",
}


def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _num(v, default=0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def main() -> None:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    path = args[0] if args else DEFAULT_PATH

    import openpyxl
    from src import database as db

    print(f"Source: {path}")
    db.init_db()

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"ERROR: sheet '{SHEET}' not found. Sheets: {wb.sheetnames}")
        return
    ws = wb[SHEET]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    idx = {h: i for i, h in enumerate(header)}

    def g(row, col):
        i = idx.get(col)
        return _s(row[i]) if i is not None and i < len(row) else ""

    print("Reading per-domain formats…")
    # email_patterns: keep the strongest sample_count per (domain, code)
    ep_seen: dict[tuple, int] = {}
    # company_email_formats: one row per (domain, format_rank) — last wins
    cef_rows: dict[tuple, dict] = {}
    total = skipped_code = 0

    for row in it:
        total += 1
        domain = g(row, "Domain").lower()
        code = g(row, "Format Code")
        if not domain or not code:
            continue

        try:
            rank = int(_num(g(row, "Format Rank"), 1))
        except ValueError:
            rank = 1

        predictable = code in _ENGINE_CODES
        if not predictable:
            skipped_code += 1

        # Per-row share % is embedded in "Multi-Format Guidance" as
        # "{code} (XX.X%, n=N)" — "Primary Share %" is only the top format's
        # share repeated on every row, not this row's own share.
        guidance = g(row, "Multi-Format Guidance")
        m = re.search(re.escape(code) + r"\s*\(([\d.]+)%", guidance)
        # "Primary Share %" is only meaningful for rank 1 (it's the same value
        # repeated on every row for a company); don't let it leak onto other
        # ranks when this row's own share can't be parsed out of the guidance.
        share_pct = float(m.group(1)) if m else (_num(g(row, "Primary Share %")) if rank == 1 else 0.0)

        format_count = int(_num(g(row, "Format Count"), 1))

        cef_rows[(domain, rank)] = {
            "company_name":       g(row, "Company"),
            "domain":             domain,
            "format_rank":        rank,
            "format_code":        code,
            "formula":            g(row, "Formula"),
            "description":        g(row, "Description"),
            "domain_example":     g(row, "Domain Example"),
            "share_pct":          share_pct,
            "format_count":       format_count,
            "sample_emails":      g(row, "Sample Emails"),
            "contacts_280k":      int(_num(g(row, "280K Contacts"))),
            "validated_emails":   int(_num(g(row, "Validated Emails"))),
            "formats_found":      int(_num(g(row, "Formats Found"), 1)),
            "is_predictable":     predictable,
            "recommended_action": g(row, "Recommended Action"),
        }

        if predictable:
            # sample_count: validated Format Count, with a small rank bias so
            # the primary format (rank 1) outsorts secondaries on ties.
            weight = format_count * 100 + max(0, 100 - rank)
            key = (domain, code)
            if weight > ep_seen.get(key, 0):
                ep_seen[key] = weight

    print(f"Parsed {total:,} rows → {len(ep_seen):,} usable email_patterns pairs, "
          f"{len(cef_rows):,} company_email_formats rows "
          f"({skipped_code:,} rows have no buildable template)")

    print("Upserting into email_patterns…")
    ep_list = [(dom, pat, cnt) for (dom, pat), cnt in ep_seen.items()]
    inserted = 0
    for i in range(0, len(ep_list), BATCH):
        chunk = ep_list[i:i + BATCH]
        with db.db_cursor() as cur:
            cur.executemany(
                # Weights are deterministic per run (ep_seen dict already took the
                # max per pair), so EXCLUDED is correct on re-runs — and portable
                # (SQLite has 2-arg MAX(), Postgres uses GREATEST(); this avoids both).
                "INSERT INTO email_patterns (domain, pattern, sample_count) VALUES (%s, %s, %s) "
                "ON CONFLICT (domain, pattern) DO UPDATE SET sample_count = EXCLUDED.sample_count",
                chunk,
            )
        inserted += len(chunk)
    print(f"  {inserted:,} email_patterns rows upserted.")

    print("Upserting into company_email_formats…")
    cef_list = list(cef_rows.values())
    upserted = 0
    for i in range(0, len(cef_list), BATCH):
        chunk = cef_list[i:i + BATCH]
        upserted += db.upsert_company_email_formats(chunk)
        if upserted % 20000 < BATCH:
            print(f"  …{upserted:,} upserted")
    print(f"  {upserted:,} company_email_formats rows upserted.")

    print("\nDONE.")
    _sample(db)


def _sample(db) -> None:
    with db.db_cursor(commit=False) as cur:
        cur.execute("SELECT COUNT(DISTINCT domain) AS d, COUNT(*) AS p FROM email_patterns")
        r = cur.fetchone()
        print(f"email_patterns covers {r['d']:,} domains ({r['p']:,} total patterns).")

    stats = db.company_email_formats_stats()
    print(f"company_email_formats covers {stats['domains']:,} domains "
          f"({stats['total_rows']:,} total rows, "
          f"{stats['predictable_domains']:,} with a buildable primary format).")


if __name__ == "__main__":
    main()
